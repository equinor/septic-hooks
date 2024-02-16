import argparse
import csv
import os
from pathlib import Path
from typing import Sequence

import openpyxl as xl


def main(argv: Sequence[str] | None = None):
    parser = argparse.ArgumentParser()
    parser.add_argument("filenames", nargs="*")
    args = parser.parse_args(argv)

    for filename in args.filenames:
        print(filename)
        output_path = Path(filename).parent / "scg_data"
        if not output_path.exists():
            os.makedirs(output_path.resolve())
        wb = xl.load_workbook(filename)
        for sheet in wb.sheetnames:
            working_sheet = wb.get_sheet_by_name(sheet)
            output_file = output_path / f"{sheet}.csv"
            with open(output_file, "w", newline="") as file_handle:
                csv_writer = csv.writer(file_handle, delimiter=";")
                for row in working_sheet.iter_rows():
                    csv_writer.writerow([cell.value for cell in row])

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
